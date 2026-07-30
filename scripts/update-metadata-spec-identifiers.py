"""Replace the old numerical requirement identifiers (e.g. 1.1, 1.7.16) in the
SAR metadata specification Excel file with the new textual identifiers
(e.g. meta-trace-sar) used in the generated PFS documents.

Usage:
    python scripts/update-metadata-spec-identifiers.py [--dry-run]

The mapping is derived from two sources:

1. The generated documents in build/ (NRB, POL, ORB, GSLC, INSAR, CB) provide
   the category-prefixed identifier (e.g. `src-acqpar`) and the building block
   YAML file it was generated from (via the `<!-- edit:... -->` comments).
2. The `history` entry of each building block YAML refers back to the original
   SAR requirement number(s) (e.g. `SAR 1.6.4`).

When an old number maps to several building blocks (product-specific variants
such as backscatter-nrb/-pol/..., or a requirement that was split), the
candidates are narrowed down using the product tags of the Excel row (e.g.
`[NRB] [POL]`) and the item name; remaining candidates are joined with ", ".

Besides plain numbers in the identifier columns, comma-separated lists
("3.1, 3.6, 3.7") and ranges ("2.2 - 2.16", expanded to all numbers in
between) are handled, as well as references to items inside the notes and
parameter text columns ("... as for 1.6.5"). For inline references the product
context is taken from the product tags of the item block the cell belongs to.

The Excel file is modified surgically: only the affected cells are rewritten
inside the worksheet XML (as inline strings, keeping the cell style), so all
other content, formatting, images etc. remain untouched. Note that some cells
store the number as a float, where e.g. 2.20 is indistinguishable from 2.2
except through the cell number format (0.00), which is taken into account.
"""

import collections
import os
import re
import shutil
import sys
import zipfile

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "CEOS-ARD_Metadata-spec_Synthetic_Aperture_Radar_v1.3.xlsx")
SAR_DOCS = ["NRB", "POL", "ORB", "GSLC", "INSAR", "CB"]

# sheet name -> (column with the old number, column with product tags or a
# fixed product set implied by the sheet, column with the item name)
SHEETS = {
    "General Metadata": (1, 3, 2),
    "Per-Pixel Metadata": (1, 3, 2),
    "Radiometric Corrections": (1, 3, 2),
    "Geometric Corrections": (1, 3, 2),
    "Change history": (2, 1, 3),
    "New parameters CB": (1, {"CB"}, 2),
    "New parameters GSLC": (1, {"GSLC"}, 2),
    "New parameters InSAR": (1, {"INSAR"}, 2),
}

PRODUCT_TAGS = {"NRB": "NRB", "POL": "POL", "ORB": "ORB", "GSLC": "GSLC", "INSAR": "InSAR", "CB": "CB"}

# sheet name -> (text columns to scan for inline item references, product tag column)
INLINE_SHEETS = {
    "General Metadata": (range(4, 9), 3),
    "Per-Pixel Metadata": (range(4, 9), 3),
    "Radiometric Corrections": (range(4, 9), 3),
    "Geometric Corrections": (range(4, 9), 3),
    "Change history": ([7], 1),
}

# grouping headers in the old spec that correspond to a whole requirement
# category (sections/requirement-categories) rather than a building block;
# they get the category id and take precedence over the history-based mapping
# (acquisition-id.yaml claims SAR 1.6, but the header row means the category)
CATEGORIES = {
    "1.6": "src",  # Source Data Attributes -> Source Metadata
    "1.7": "prd",  # CEOS-ARD Product Attributes -> Product Metadata
}

# manual corrections of typos in the source file: (sheet, cell, number) -> number
NUMBER_OVERRIDES = {
    # "dB = 10*log10(DN) where DN is 2.2" inside item 2.22 (Composite Quality
    # Map): DN is the quality map value itself, i.e. 2.22 (cf. the analogous
    # "DN is 3.1" notes in the Radiometric Corrections sheet)
    ("Per-Pixel Metadata", "G216", "2.2"): "2.22",
}

NUMBER_RE = re.compile(r"(?<![\dv.:])\d+(?:\.\d+)+(?![\d.])")


def build_mapping():
    """number -> list of (identifier, title), plus identifier -> set of docs"""
    id_docs = collections.defaultdict(set)
    block_ids = collections.defaultdict(set)
    for doc in SAR_DOCS:
        with open(os.path.join(ROOT, "build", f"{doc}.md"), encoding="utf-8") as f:
            text = f.read()
        for m in re.finditer(r"<!-- edit:([^>]+?)-->", text):
            path = m.group(1).replace("\\", "/")
            idm = re.search(r"Identifier: `([a-z0-9-]+)`", text[m.end():m.end() + 600])
            if idm and "requirements/" in path:
                rel = path.split("requirements/")[1]
                block_ids[rel].add(idm.group(1))
                id_docs[idm.group(1)].add(doc)

    num_map = collections.defaultdict(set)
    for rel, ids in block_ids.items():
        with open(os.path.join(ROOT, "requirements", rel), encoding="utf-8") as f:
            y = f.read()
        tm = re.search(r"^title:\s*(.+)", y, re.M)
        title = tm.group(1).strip() if tm else ""
        hm = re.search(r"^history:\s*\n((?:[ \t]+-[ \t]+.+\n?)*)", y, re.M)
        if not hm:
            continue
        for line in hm.group(1).strip().splitlines():
            line = line.strip().lstrip("- ").strip()
            # only SAR history entries; ignore product qualifiers in parentheses
            base = line.split("(")[0]
            if "SAR" not in base:
                continue
            for num in re.findall(r"(\d+(?:\.\d+)+)", base):
                for i in ids:
                    num_map[num].add((i, title))
    return num_map, id_docs


def products_from_tags(text):
    """Parse '[NRB] [POL]' style product tags; None means no restriction."""
    tags = set(re.findall(r"\[([A-Za-z]+)\]", text or ""))
    if "ALL" in tags:
        return None
    prods = {doc for doc, tag in PRODUCT_TAGS.items() if tag in tags or tag.upper() in tags}
    return prods or None


def resolve(num, products, item_name, num_map, id_docs):
    cands = num_map.get(num)
    if not cands:
        return None
    cands = set(cands)

    def ids(c):
        return {i for i, _ in c}

    # a block can appear under both a general-metadata and a per-pixel
    # identifier (e.g. meta-memare-sar / pxl-memare-sar); the chapter number
    # decides which one applies
    if len(ids(cands)) > 1:
        pxl = {c for c in cands if c[0].startswith("pxl-")}
        preferred = pxl if num.startswith("2.") else cands - pxl
        if preferred:
            cands = preferred
    # narrow product-specific variants down via the row's product tags
    if products and len(ids(cands)) > 1:
        filtered = {c for c in cands if id_docs[c[0]] & products}
        if filtered:
            cands = filtered
    # narrow split requirements down via the item name
    if item_name and len(ids(cands)) > 1:
        norm = lambda s: re.sub(r"\W+", "", s.lower())
        matched = {c for c in cands if norm(c[1]) in norm(item_name) or norm(item_name) in norm(c[1])}
        if matched and len(ids(matched)) < len(ids(cands)):
            cands = matched
    return sorted(ids(cands))


def formatted_number(cell):
    """The number as displayed, respecting the number format (2.20 vs 2.2)."""
    v = cell.value
    if isinstance(v, float):
        m = re.search(r"0\.(0+)", cell.number_format)
        if m:
            return f"{v:.{len(m.group(1))}f}"
        s = repr(v)
        return s.rstrip("0").rstrip(".") if "." in s else s
    return str(v).strip()


def expand_numbers(text):
    """Parse an identifier cell into a list of numbers: a plain number, a
    comma-separated list, or a range ("2.2 - 2.16", expanded)."""
    range_m = re.fullmatch(r"(\d+(?:\.\d+)+)\s*[-–—]\s*(\d+(?:\.\d+)+)", text)
    if range_m:
        start, end = (n.split(".") for n in range_m.groups())
        if start[:-1] == end[:-1] and int(start[-1]) < int(end[-1]):
            prefix = ".".join(start[:-1])
            return [f"{prefix}.{i}" for i in range(int(start[-1]), int(end[-1]) + 1)]
    tokens = [t.strip() for t in text.split(",")]
    if all(re.fullmatch(r"\d+(\.\d+)+", t) for t in tokens):
        return tokens
    return None


def block_products(ws, row, prod_col):
    """Product tags of the item block a row belongs to (nearest tag cell at or
    above the row)."""
    for r in range(row, 4, -1):
        v = ws.cell(r, prod_col).value
        if isinstance(v, str) and re.search(r"\[[A-Za-z]+\]", v):
            return products_from_tags(v)
    return None


def compute_plan():
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    plan = {}
    unmapped = []
    for name, (id_col, prod_src, name_col) in SHEETS.items():
        ws = wb[name]
        changes = {}
        for row in ws.iter_rows(min_col=id_col, max_col=id_col):
            cell = row[0]
            if cell.row <= 4 or cell.value is None:
                continue
            num = formatted_number(cell)
            nums = expand_numbers(num)
            if not nums:
                continue
            if isinstance(prod_src, set):
                products = prod_src
            else:
                products = products_from_tags(ws.cell(cell.row, prod_src).value if isinstance(ws.cell(cell.row, prod_src).value, str) else None)
            item_name = ws.cell(cell.row, name_col).value
            results = []
            for n in nums:
                if n in CATEGORIES:
                    results.append(CATEGORIES[n])
                    continue
                result = resolve(n, products, item_name if isinstance(item_name, str) and len(nums) == 1 else None, NUM_MAP, ID_DOCS)
                if result is None:
                    unmapped.append((name, cell.coordinate, n, item_name))
                else:
                    results.append(", ".join(result))
            if results and len(results) == len(nums):
                changes[cell.coordinate] = (num, ", ".join(results))
        plan[name] = changes

    # inline references to items in notes/parameter text columns
    for name, (cols, prod_col) in INLINE_SHEETS.items():
        ws = wb[name]
        changes = plan[name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if cell.row <= 4 or cell.column not in cols or not isinstance(v, str):
                    continue
                if "http" in v or "version=" in v or not NUMBER_RE.search(v):
                    continue

                def repl(m):
                    num = NUMBER_OVERRIDES.get((name, cell.coordinate, m.group(0)), m.group(0))
                    result = resolve(num, block_products(ws, cell.row, prod_col), None, NUM_MAP, ID_DOCS)
                    return ", ".join(result) if result else m.group(0)

                new = NUMBER_RE.sub(repl, v)
                if new != v:
                    changes[cell.coordinate] = (v, new)
    return plan, unmapped


def apply_plan(plan):
    """Rewrite only the affected cells inside the worksheet XML files."""
    with zipfile.ZipFile(XLSX) as z:
        workbook = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rel_target = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels))
        sheet_file = {}
        for m in re.finditer(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', workbook):
            target = rel_target[m.group(2)].lstrip("/")
            sheet_file[m.group(1)] = target if target.startswith("xl/") else "xl/" + target
        items = [(i, z.read(i.filename)) for i in z.infolist()]

    file_changes = {sheet_file[name]: ch for name, ch in plan.items() if ch}
    tmp = XLSX + ".tmp"
    replaced = 0
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for info, data in items:
            if info.filename in file_changes:
                xml = data.decode("utf-8")
                for coord, (_, new) in file_changes[info.filename].items():
                    escaped = new.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    def repl(m, new=escaped):
                        style = re.search(r'\ss="\d+"', m.group(0))
                        s_attr = style.group(0) if style else ""
                        return f'<c r="{m.group(1)}"{s_attr} t="inlineStr"><is><t xml:space="preserve">{new}</t></is></c>'
                    xml, n = re.subn(rf'<c r="({coord})"(?=[ />])[^>]*(?:/>|>.*?</c>)', repl, xml)
                    if n != 1:
                        raise RuntimeError(f"cell {coord} matched {n} times in {info.filename}")
                    replaced += 1
                data = xml.encode("utf-8")
            out.writestr(info, data)
    shutil.move(tmp, XLSX)
    return replaced


if __name__ == "__main__":
    NUM_MAP, ID_DOCS = build_mapping()
    plan, unmapped = compute_plan()

    for name, changes in plan.items():
        print(f"== {name} ({len(changes)} cells)")
        for coord, (old, new) in changes.items():
            print(f"  {coord}: {old} -> {new}")
    if unmapped:
        print("\nUnmapped (left unchanged):")
        for name, coord, num, item in unmapped:
            print(f"  {name}!{coord}: {num} ({item})")

    if "--dry-run" in sys.argv:
        print("\nDry run, file not modified.")
    else:
        replaced = apply_plan(plan)
        print(f"\nReplaced {replaced} cells in {os.path.basename(XLSX)}")
